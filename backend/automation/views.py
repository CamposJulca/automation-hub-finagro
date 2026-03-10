import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from rest_framework import viewsets, filters
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from rest_framework import status
from .models import Automation, NecesidadInnovacion
from .serializers import AutomationSerializer, NecesidadInnovacionSerializer

TIPO_VALOR = {'Operativo': 2, 'Normativo': 3, 'Estrategico': 1}


def calcular_scores(necesidades):
    """
    Fórmula del Excel:
      score = (rec / max_rec) * 0.30
            + (dur / max_dur) * 0.25
            + (riesgo / 3)   * 0.30
            + (tipo_val / 3) * 0.15
    max_rec y max_dur son dinámicos sobre el dataset.
    """
    if not necesidades:
        return []
    max_rec = max(n.recurrencia_mensual for n in necesidades) or 1
    max_dur = max(float(n.duracion_horas) for n in necesidades) or 1
    result = []
    for n in necesidades:
        tipo_val = TIPO_VALOR.get(n.tipo_objetivo, 1)
        score = (
            (n.recurrencia_mensual / max_rec) * 0.30 +
            (float(n.duracion_horas) / max_dur) * 0.25 +
            (n.riesgo / 3) * 0.30 +
            (tipo_val / 3) * 0.15
        )
        result.append((n, round(score, 4)))
    result.sort(key=lambda x: x[1], reverse=True)
    return result


class AutomationViewSet(viewsets.ModelViewSet):
    queryset = Automation.objects.all()
    serializer_class = AutomationSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter]
    search_fields = ['name', 'module', 'description']

    def get_queryset(self):
        qs = super().get_queryset()
        module = self.request.query_params.get('module')
        active = self.request.query_params.get('active')
        if module:
            qs = qs.filter(module=module)
        if active is not None:
            qs = qs.filter(active=active.lower() == 'true')
        return qs


class NecesidadInnovacionViewSet(viewsets.ModelViewSet):
    queryset = NecesidadInnovacion.objects.all()
    serializer_class = NecesidadInnovacionSerializer
    permission_classes = [AllowAny]

    def list(self, request, *args, **kwargs):
        necesidades = list(self.get_queryset())
        scored = calcular_scores(necesidades)
        data = []
        for necesidad, score in scored:
            s = NecesidadInnovacionSerializer(necesidad)
            item = s.data
            item['puntuacion'] = score
            data.append(item)
        return Response(data)

    def create(self, request, *args, **kwargs):
        response = super().create(request, *args, **kwargs)
        necesidades = list(NecesidadInnovacion.objects.all())
        scored = calcular_scores(necesidades)
        data = []
        for necesidad, score in scored:
            s = NecesidadInnovacionSerializer(necesidad)
            item = s.data
            item['puntuacion'] = score
            data.append(item)
        return Response(data, status=response.status_code)

    @action(detail=False, methods=['post'], url_path='importar', permission_classes=[AllowAny])
    def importar(self, request):
        """Importa necesidades desde un Excel con el formato de la Matriz de Innovación."""
        archivo = request.FILES.get('archivo')
        if not archivo:
            return Response({'error': 'No se recibió ningún archivo.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            wb = openpyxl.load_workbook(archivo, data_only=True)
            ws = wb['Procesos manuales']
        except Exception:
            return Response({'error': 'El archivo no es válido o no contiene la hoja "Procesos manuales".'}, status=status.HTTP_400_BAD_REQUEST)

        TIPOS_VALIDOS = {'Operativo', 'Normativo', 'Estrategico', 'Estratégico'}
        creados = 0
        errores = []

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Columnas: A(idx0)=None, B=Área, C=Proceso, D=Responsable,
            #           E=Situación, F=Preocupaciones, G=Necesidad,
            #           H=Tipo, I=Recurrencia, J=Duración, K=Riesgo, L=Puntuación
            if len(row) < 11:
                continue
            _, area, proceso, responsable, situacion, preocupaciones, necesidad, tipo, recurrencia, duracion, riesgo = row[:11]

            # Saltar filas completamente vacías o que sean la fila de cabeceras
            if all(v is None for v in [area, proceso, necesidad, tipo, recurrencia, duracion, riesgo]):
                continue
            if str(tipo or '').strip() in ('Tipo de objetivo', 'TipoProceso', 'Tipo'):
                continue

            # Normalizar tipo
            if tipo == 'Estratégico':
                tipo = 'Estrategico'

            # Validaciones mínimas
            if tipo not in TIPOS_VALIDOS:
                errores.append(f'Fila {i}: tipo "{tipo}" no reconocido.')
                continue
            try:
                recurrencia = int(recurrencia)
                duracion    = float(duracion)
                riesgo      = int(riesgo)
                assert 1 <= riesgo <= 3
                assert recurrencia >= 1
                assert duracion > 0
            except Exception:
                errores.append(f'Fila {i}: valores numéricos inválidos.')
                continue

            NecesidadInnovacion.objects.create(
                area              = str(area or ''),
                proceso_asociado  = str(proceso or ''),
                responsable       = str(responsable or ''),
                situacion_actual  = str(situacion or ''),
                preocupaciones    = str(preocupaciones or ''),
                necesidad         = str(necesidad or ''),
                tipo_objetivo     = tipo,
                recurrencia_mensual = recurrencia,
                duracion_horas    = duracion,
                riesgo            = riesgo,
            )
            creados += 1

        necesidades = list(NecesidadInnovacion.objects.all())
        scored = calcular_scores(necesidades)
        data = []
        for necesidad_obj, score in scored:
            item = NecesidadInnovacionSerializer(necesidad_obj).data
            item['puntuacion'] = score
            data.append(item)

        return Response({'creados': creados, 'errores': errores, 'necesidades': data}, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['get'], url_path='exportar', permission_classes=[AllowAny])
    def exportar(self, request):
        """Exporta todas las necesidades priorizadas a un Excel."""
        necesidades = list(NecesidadInnovacion.objects.all())
        scored = calcular_scores(necesidades)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Priorización Innovación'

        # Estilos
        verde    = '004d26'
        dorado   = 'c9a227'
        gris_bg  = 'f2f4f3'

        hdr_font  = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        hdr_fill  = PatternFill('solid', fgColor=verde)
        hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin      = Side(style='thin', color='D0DBD5')
        border    = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers = [
            '#', 'Prioridad', 'Score (%)', 'Área', 'Proceso Asociado',
            'Responsable', 'Necesidad Identificada', 'Situación Actual',
            'Preocupaciones', 'Tipo de Objetivo',
            'Recurrencia Mensual', 'Duración (h)', 'Riesgo',
        ]
        col_widths = [5, 10, 11, 22, 26, 24, 40, 36, 36, 14, 14, 11, 10]

        for col, (header, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font  = hdr_font
            cell.fill  = hdr_fill
            cell.alignment = hdr_align
            cell.border = border
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.row_dimensions[1].height = 32

        PRIORIDAD = {(0.70, 1.01): 'Alta', (0.50, 0.70): 'Media', (0.0, 0.50): 'Baja'}
        RIESGO_LABEL = {1: 'Bajo', 2: 'Medio', 3: 'Alto'}
        COLOR_ALTA  = 'FFEBEE'
        COLOR_MEDIA = 'FFF3E0'
        COLOR_BAJA  = 'E8F5E9'

        for row_idx, (n, score) in enumerate(scored, start=2):
            if score >= 0.70:
                prioridad, row_color = 'Alta',  COLOR_ALTA
            elif score >= 0.50:
                prioridad, row_color = 'Media', COLOR_MEDIA
            else:
                prioridad, row_color = 'Baja',  COLOR_BAJA

            row_fill  = PatternFill('solid', fgColor=row_color)
            score_pct = round(score * 100, 2)

            valores = [
                row_idx - 1, prioridad, score_pct,
                n.area, n.proceso_asociado, n.responsable,
                n.necesidad, n.situacion_actual, n.preocupaciones,
                n.tipo_objetivo, n.recurrencia_mensual,
                float(n.duracion_horas), RIESGO_LABEL.get(n.riesgo, n.riesgo),
            ]

            for col, valor in enumerate(valores, start=1):
                cell = ws.cell(row=row_idx, column=col, value=valor)
                cell.fill   = row_fill
                cell.border = border
                cell.alignment = Alignment(vertical='center', wrap_text=(col >= 7))
                if col == 2:  # Prioridad — negrita
                    cell.font = Font(bold=True)
                if col == 3:  # Score — negrita + color
                    cell.font = Font(bold=True, color=verde)

            ws.row_dimensions[row_idx].height = 28

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'

        # Hoja de fórmula
        ws2 = wb.create_sheet('Fórmula')
        ws2['B2'] = 'Fórmula de priorización'
        ws2['B2'].font = Font(bold=True, size=12, color=verde)
        params = [
            ('Recurrencia', '30%', 'rec / max(recurrencias)'),
            ('Duración',    '25%', 'dur / max(duraciones)'),
            ('Riesgo',      '30%', 'riesgo / 3  (escala 1-3)'),
            ('Tipo proceso','15%', 'Normativo=3 · Operativo=2 · Estratégico=1  / 3'),
        ]
        ws2['B3'] = 'Score = ' + ' + '.join(f'({p[0]} / máx) × {p[1]}' for p in params)
        for r, (factor, peso, norm) in enumerate(params, start=5):
            ws2.cell(row=r, column=2, value=factor).font = Font(bold=True)
            ws2.cell(row=r, column=3, value=peso)
            ws2.cell(row=r, column=4, value=norm)
        ws2.column_dimensions['B'].width = 18
        ws2.column_dimensions['C'].width = 8
        ws2.column_dimensions['D'].width = 44

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="Priorizacion_Innovacion.xlsx"'
        return response
